import csv
import openpyxl
import cgi
from flask import Flask, flash, jsonify, render_template, request
import webbrowser
app = Flask(__name__)

form = cgi.FieldStorage()

data_file = "./zomato.csv"
country_codes_file = "./Country-Code.xlsx"

def read_data_file():
    column_names = []
    data = []
    
    with open(data_file, 'r', encoding='UTF-8', newline='', errors='ignore') as f:
        csv_reader = csv.reader(f)
        first_row = True
        for row in csv_reader:
            if first_row:
                column_names = row
                first_row = False
                continue
            current_data = {}
            for i, cell in enumerate(row):
                current_data[column_names[i]] = str(cell)
            data.append(current_data)
    return column_names, data

def read_country_codes():
    wb = openpyxl.load_workbook(country_codes_file, data_only=True)
    ws = wb['Sheet1']

    country_codes = {}
    
    for i in range(2, ws.max_row + 1):
        for j in range(1, 2, ws.max_column + 1):
            country_codes[ws.cell(row=i, column=j).value] = ws.cell(row=i, column=j+1).value
    
    return country_codes

@app.route('/', methods=['GET', 'POST'])
def index():
    return render_template('index.html')

@app.route('/data', methods=['POST'])
def data():
    columns, data = read_data_file()
    country_codes = read_country_codes()
    cuisines = request.form['cuisine']
    country = request.form['country']
    city = request.form['city']
    currency = request.form['currency']
    message = ""

    total = dict()

    top_list = dict()

    for d in data:
        current_cuisines = dict()
        for cuisine in cuisines.split(", "):
            current_cuisines[cuisine] = 0
            if cuisine in d['Cuisines']:
                current_cuisines[cuisine] += 1

        current_country = 0
        if country_codes[int(d["Country Code"])] == country:
            current_country += 1

        current_city = 0
        if city == d["City"]:
            current_city += 1

        current_currency = 0
        if currency == d["Currency"]:
            current_currency += 1

        top_list[d["Restaurant Name"]] = sum(current_cuisines.values())/len(d['Cuisines'].split(', ')) + current_country + current_city + current_currency

    for key, value in {k: v for k, v in sorted(top_list.items(), key=lambda item: item[1], reverse=True)}.items():
        message += "{} - {}\n".format(key, str(value))

    return render_template('data.html', message=message)

app.run(host='localhost', port=5000)